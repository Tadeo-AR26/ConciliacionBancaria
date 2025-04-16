import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

def generar_discrepancias(df_ext, df_cont, output_path):
    # Agrupar por color y sumar valores absolutos
    resumen_ext = df_ext[df_ext['_color'] != ""]\
        .groupby('_color')['Importe Pesos'].apply(lambda x: x.abs().sum())

    resumen_cont = df_cont[df_cont['_color'] != ""]\
        .groupby('_color')['Movimientos'].apply(lambda x: x.abs().sum())

    # Colores combinados
    colores = set(resumen_ext.index).union(resumen_cont.index)

    wb = Workbook()
    wb.remove(wb.active)

    for i, color in enumerate(colores, start=1):
        total_ext = resumen_ext.get(color, 0)
        total_cont = resumen_cont.get(color, 0)

        if round(total_ext, 2) != round(total_cont, 2):
            hoja = wb.create_sheet(f"Discrepancia {i}")

            # Filas de extracto
            filas_ext = df_ext[df_ext['_color'] == color][['Cod. Operativo', 'Importe Pesos']].copy()
            filas_ext.rename(columns={'Cod. Operativo': 'Codigo Operativo'}, inplace=True)
            filas_ext['Fuente'] = 'Extracto'

            # Filas de contabilidad
            filas_cont = df_cont[df_cont['_color'] == color][['Observ.', 'Movimientos']].copy()
            filas_cont.rename(columns={'Observ.': 'Observacion'}, inplace=True)
            filas_cont['Fuente'] = 'Contabilidad'

            # Normalizamos columnas para concatenar
            filas_ext['Observacion'] = ""
            filas_cont['Codigo Operativo'] = ""

            # Reordenar para consistencia
            columnas = ['Codigo Operativo', 'Importe Pesos', 'Observacion', 'Movimientos', 'Fuente']
            df_resultado = pd.concat([filas_ext, filas_cont], ignore_index=True)[columnas]

            for r_idx, row in enumerate(dataframe_to_rows(df_resultado, index=False, header=True), 1):
                for c_idx, val in enumerate(row, 1):
                    cell = hoja.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx > 1:  # Saltamos la cabecera
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            # Escribir diferencia total al final
            hoja.append([])
            hoja.append(["", "", "", "", "Diferencia total:"])
            hoja.append(["", "", "", "", round(abs(total_ext - total_cont), 2)])

    wb.save(output_path)
