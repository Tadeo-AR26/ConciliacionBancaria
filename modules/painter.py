import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

def pintar_extracto(df, referencias, nombre_archivo="extracto_coloreado.xlsx"):
    codigos = referencias['codigo']
    df['_color'] = df['Cod. Operativo'].map(codigos).fillna("")

    # Guardamos a Excel primero
    df.to_excel(nombre_archivo, index=False)

    # Aplicamos color a filas en el archivo
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    for idx, color in enumerate(df['_color'], start=2):  # empieza en 2 por el header
        if color:
            fill = PatternFill(start_color=color.replace("#", ""), end_color=color.replace("#", ""), fill_type="solid")
            for cell in ws[idx]:
                cell.fill = fill

    wb.save(nombre_archivo)
    return df


def pintar_contabilidad(df, referencias, nombre_archivo="contabilidad_coloreado.xlsx"):
    # Convertimos los textos del diccionario de observaciones en expresiones regulares flexibles
    patrones = []
    for clave, color in referencias['observacion'].items():
        # Escapamos cualquier carácter especial en la clave y la convertimos en una regex tipo .*clave.*
        patron_regex = re.compile(rf".*{re.escape(clave)}.*", re.IGNORECASE)
        patrones.append((patron_regex, color))

    colores = []
    for observ in df['Observ.'].astype(str):
        color = ""
        for regex, c in patrones:
            if regex.search(observ):
                color = c
                break
        colores.append(color)

    df['_color'] = colores

    # Guardamos a Excel
    df.to_excel(nombre_archivo, index=False)

    # Aplicamos color a filas en el archivo
    wb = load_workbook(nombre_archivo)
    ws = wb.active

    for idx, color in enumerate(df['_color'], start=2):  # empieza en 2 por el header
        if color:
            fill = PatternFill(start_color=color.replace("#", ""), end_color=color.replace("#", ""), fill_type="solid")
            for cell in ws[idx]:
                cell.fill = fill

    wb.save(nombre_archivo)
    return df